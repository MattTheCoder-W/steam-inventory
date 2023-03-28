import os
import requests
import json
import openpyxl
from html import unescape
import time
from bs4 import BeautifulSoup as bs


def check_prices(filepath):
    if not os.path.exists(filepath):
        return "File not found!"
    
    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]
    new_ws = wb.copy_worksheet(ws)
    new_ws.title = "Updated"

    for row in range(1, ws.max_row + 1):
        name_address = f"A{row}"
        price_address = f"C{row}"
        name = ws[name_address].value
        price = ws[price_address].value
        if not price or not name:
            continue
        if "=" in str(price) or "=" in str(name):
            continue
        print(name, price)
        html_name = unescape(name)
        url = f"https://steamcommunity.com/market/search?q={html_name}&appid=730"
        page = requests.get(url).content
        soup = bs(page, 'html.parser')
        results = soup.select("div#searchResultsRows")
        if not results:
            print("Items not found ->", name)
        try:
            results = results[0]
            item = results.select("a")[0]
            price = item.select("span.normal_price[data-price]")[0]
            price = int(price['data-price'])/100
            print("Price in USD:", price)
        except:
            print(f"Error occurred -> {name}")
            continue
        new_ws[price_address].value = price * 4.32

    wb.save("new.xlsx")

    return "Scraping file from python!"
