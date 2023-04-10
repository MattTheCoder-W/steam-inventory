import openpyxl
import os


class Loader:
    _input_path = None
    def __init__(self, input_path: str):
        self.data = []
        self.input_path = input_path
        self.load_data()

    @property
    def input_path(self) -> str:
        return self._input_path
    
    @input_path.setter
    def input_path(self, value: str):
        if not os.path.exists(value):
            raise FileNotFoundError(f"Cannot locate inpurt data file -> {value}")
        self._input_path = value
    
    def load_data(self):
        wb = openpyxl.load_workbook(self.input_path)
        ws = wb[wb.sheetnames[0]]

        for row in range(2, ws.max_row + 1):
            name_address = f"A{row}"
            item_type_address = f"B{row}"
            price_address = f"D{row}"
            url_address = f"H{row}"
            last_address = f"I{row}"
            name = ws[name_address].value
            item_type = ws[item_type_address].value
            price = ws[price_address].value
            url = ws[url_address].value
            url = "" if url is None else url
            
            if name is None or price is None or (isinstance(price, str) and "=" in price):
                continue

            self.data.append({
                "addresses": {
                    "name_address": name_address,
                    "price_address": price_address,
                    "url_address": url_address,
                    "last_price_address": last_address,
                },
                "name": name, 
                "type": item_type,
                "price": price,
                "last_price": "",
                "url": url,
            })

    def save_urls(self, filename):
        wb = openpyxl.load_workbook(self.input_path)
        ws = wb[wb.sheetnames[0]]

        for item in self.data:
            ws[item['addresses']['url_address']].value = item['url']
        
        wb.save(filename)
        print("[info] output file saved!")
    
    def save_prices(self, filename):
        wb = openpyxl.load_workbook(self.input_path)
        ws = wb[wb.sheetnames[0]]

        for item in self.data:
            ws[item['addresses']['last_price_address']].value = item['last_price']
        
        wb.save(filename)
        print("[info] output file saved!")
    

